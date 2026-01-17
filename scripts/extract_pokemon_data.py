#!/usr/bin/env python3
"""
extract_pokemon_data.py
Extrahiert Pokemon-Daten aus Excel und speichert alle 8 Generationen als JSON.

Verwendung:
    python extract_pokemon_data.py
"""

import json
import openpyxl
import sys
import re
from pathlib import Path

# Generation ID-Bereiche
GENERATION_RANGES = {
    1: (1, 151, 'Kanto'),
    2: (152, 251, 'Johto'),
    3: (252, 386, 'Hoenn'),
    4: (387, 493, 'Sinnoh'),
    5: (494, 649, 'Unova'),
    6: (650, 721, 'Kalos'),
    7: (722, 809, 'Alola'),
    8: (810, 905, 'Galar'),
}


def extract_image_url(formula_str):
    """Extrahiert die URL aus einer Excel-Image-Formel."""
    if not formula_str or not isinstance(formula_str, str):
        return None
    
    match = re.search(r'http[s]?://[^\s"\')\]]+', formula_str)
    if match:
        return match.group(0)
    return None


def extract_pokemon_number(num_str):
    """Extrahiert die Nummer aus Format wie '#001'."""
    if not num_str:
        return None
    
    num_str = str(num_str).strip()
    if num_str.startswith('#'):
        try:
            return int(num_str[1:])
        except ValueError:
            return None
    return None


def extract_generation_data(generation):
    """
    Liest Excel-Datei und extrahiert eine bestimmte Generation.
    """
    start_id, end_id, region_name = GENERATION_RANGES[generation]
    
    # Pfade
    script_dir = Path(__file__).parent
    project_dir = script_dir.parent
    excel_file = project_dir / "_greenie (Pokédex).xlsx"
    data_dir = project_dir / "data"
    
    # Verzeichnis erstellen
    data_dir.mkdir(exist_ok=True)
    
    print("=" * 80)
    print(f"Pokemon Gen {generation} ({region_name}) Datenextraktion")
    print(f"Pokemon #{start_id:03d} - #{end_id:03d}")
    print("=" * 80)
    
    # Excel öffnen
    if not excel_file.exists():
        print(f"❌ Excel-Datei nicht gefunden: {excel_file}")
        return False, None
    
    print(f"\n📂 Excel-Datei: {excel_file.name}")
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        print(f"📄 Blatt: {ws.title}")
        
        # Pokemon sammeln (mit Duplikat-Deduplication)
        pokemon_dict = {}  # ID → Pokemon
        
        # Durch alle Reihen gehen
        for row_idx in range(1, ws.max_row + 1):
            # Spalten auslesen
            num_cell = ws.cell(row_idx, 1).value
            name_en_cell = ws.cell(row_idx, 4).value
            image_cell = ws.cell(row_idx, 3).value
            type1_cell = ws.cell(row_idx, 5).value
            type2_cell = ws.cell(row_idx, 6).value
            
            # Nummer extrahieren
            pokemon_num = extract_pokemon_number(num_cell)
            
            # Nur die gewünschte Generation
            if pokemon_num is None or pokemon_num < start_id or pokemon_num > end_id:
                continue
            
            # Nur ERSTE Vorkommen behalten (Duplikate überspringen)
            if pokemon_num in pokemon_dict:
                continue
            
            # Bild-URL extrahieren
            image_url = extract_image_url(image_cell)
            
            # Daten zusammenstellen
            pokemon = {
                "id": pokemon_num,
                "num": f"#{pokemon_num:03d}",
                "name_en": str(name_en_cell).strip() if name_en_cell else None,
                "type1": str(type1_cell).strip() if type1_cell else None,
                "type2": str(type2_cell).strip() if type2_cell else None,
                "image_url": image_url,
                "generation": generation
            }
            
            pokemon_dict[pokemon_num] = pokemon
        
        # In Liste konvertieren und sortieren
        pokemon_data = sorted(pokemon_dict.values(), key=lambda x: x["id"])
        
        print(f"\n✅ {len(pokemon_data)} einzigartige Pokemon aus Gen {generation} extrahiert")
        
        # Statistik
        if pokemon_data:
            print("\n📊 Statistik:")
            print(f"  ID-Bereich:    #{pokemon_data[0]['id']:03d} - #{pokemon_data[-1]['id']:03d}")
            with_images = sum(1 for p in pokemon_data if p['image_url'])
            with_type2 = sum(1 for p in pokemon_data if p['type2'])
            print(f"  Mit Bild-URL:  {with_images}/{len(pokemon_data)}")
            print(f"  Mit Typ 2:     {with_type2}/{len(pokemon_data)}")
        
        output_file = data_dir / f"pokemon_gen{generation}_raw.json"
        
        # JSON speichern
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(pokemon_data, f, indent=2, ensure_ascii=False)
        
        print(f"\n💾 Gespeichert: {output_file.name}")
        
        return True, len(pokemon_data)
        
    except Exception as e:
        print(f"\n❌ Fehler: {e}")
        import traceback
        traceback.print_exc()
        return False, None


if __name__ == "__main__":
    print("=" * 80)
    print("Pokemon Datenextraktion - Alle Generationen")
    print("=" * 80 + "\n")
    
    total_extracted = 0
    all_success = True
    
    for gen in sorted(GENERATION_RANGES.keys()):
        success, count = extract_generation_data(gen)
        if success and count:
            total_extracted += count
        else:
            all_success = False
        print()
    
    print("=" * 80)
    print(f"Gesamt: {total_extracted} Pokemon extrahiert")
    print("=" * 80)
    
    sys.exit(0 if all_success else 1)
